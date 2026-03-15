"""
Backtesting Module

Provides comprehensive backtesting functionality for investment strategies:
- Historical portfolio simulation
- Performance metrics (Sharpe, Sortino, max drawdown, etc.)
- Benchmark comparison (S&P 500)
- Trade-by-trade analysis
- Visualizations ready data

Example:
    >>> from stock_src.backtester import Backtester, BacktestConfig
    >>> config = BacktestConfig(
    ...     start_date="2020-01-01",
    ...     end_date="2023-12-31",
    ...     initial_capital=100000,
    ...     rebalance_frequency="monthly"
    ... )
    >>> backtester = Backtester(config)
    >>> results = backtester.run_backtest(["AAPL", "MSFT", "GOOGL"])
    >>> print(f"Total Return: {results.total_return:.2%}")
"""

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import yfinance as yf

from stock_src.strategies import InvestmentStrategies


class RebalanceFrequency(Enum):
    """Rebalancing frequency options"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"


@dataclass
class BacktestConfig:
    """Configuration for backtesting"""
    start_date: str
    end_date: str
    initial_capital: float = 100000.0
    rebalance_frequency: str = "monthly"
    transaction_cost_pct: float = 0.001  # 0.1% per trade
    position_size_pct: float = 0.10  # 10% max per position
    min_score_threshold: int = 50  # Minimum total score to buy
    benchmark: str = "SPY"  # Benchmark ticker
    
    def __post_init__(self):
        """Validate configuration"""
        self.rebalance_frequency = RebalanceFrequency(self.rebalance_frequency)
        start = datetime.strptime(self.start_date, "%Y-%m-%d")
        end = datetime.strptime(self.end_date, "%Y-%m-%d")
        if start >= end:
            raise ValueError("start_date must be before end_date")


@dataclass
class Trade:
    """Represents a single trade"""
    date: datetime
    ticker: str
    action: str  # "BUY" or "SELL"
    shares: int
    price: float
    value: float
    transaction_cost: float
    reason: str = ""


@dataclass
class Position:
    """Represents a stock position"""
    ticker: str
    shares: int
    avg_cost: float
    current_price: float = 0.0
    
    @property
    def market_value(self) -> float:
        return self.shares * self.current_price
    
    @property
    def unrealized_pnl(self) -> float:
        return (self.current_price - self.avg_cost) * self.shares
    
    @property
    def unrealized_pnl_pct(self) -> float:
        if self.avg_cost == 0:
            return 0.0
        return (self.current_price / self.avg_cost - 1) * 100


@dataclass
class BacktestResults:
    """Results from a backtest run"""
    # Configuration
    config: BacktestConfig
    tickers: list
    
    # Performance metrics
    total_return: float = 0.0
    annualized_return: float = 0.0
    benchmark_return: float = 0.0
    excess_return: float = 0.0
    
    # Risk metrics
    volatility: float = 0.0
    sharpe_ratio: float = 0.0
    sortino_ratio: float = 0.0
    max_drawdown: float = 0.0
    max_drawdown_duration: int = 0
    beta: float = 0.0
    alpha: float = 0.0
    
    # Trading statistics
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    avg_trade_return: float = 0.0
    total_transaction_costs: float = 0.0
    
    # Time series data
    portfolio_values: pd.DataFrame = field(default_factory=pd.DataFrame)
    daily_returns: pd.Series = field(default_factory=pd.Series)
    benchmark_returns: pd.Series = field(default_factory=pd.Series)
    trades: list = field(default_factory=list)
    
    # Summary
    final_capital: float = 0.0
    start_date: str = ""
    end_date: str = ""
    
    def to_dict(self) -> dict:
        """Convert results to dictionary"""
        return {
            "total_return": self.total_return,
            "annualized_return": self.annualized_return,
            "benchmark_return": self.benchmark_return,
            "excess_return": self.excess_return,
            "volatility": self.volatility,
            "sharpe_ratio": self.sharpe_ratio,
            "sortino_ratio": self.sortino_ratio,
            "max_drawdown": self.max_drawdown,
            "max_drawdown_duration_days": self.max_drawdown_duration,
            "beta": self.beta,
            "alpha": self.alpha,
            "total_trades": self.total_trades,
            "winning_trades": self.winning_trades,
            "losing_trades": self.losing_trades,
            "win_rate": self.win_rate,
            "avg_trade_return": self.avg_trade_return,
            "total_transaction_costs": self.total_transaction_costs,
            "final_capital": self.final_capital,
            "initial_capital": self.config.initial_capital,
            "start_date": self.start_date,
            "end_date": self.end_date,
        }


class Backtester:
    """
    Backtests investment strategies on historical data.
    
    Features:
        - Multi-strategy scoring system
        - Configurable rebalancing
        - Transaction cost modeling
        - Risk-adjusted performance metrics
        - Benchmark comparison
    
    Example:
        >>> config = BacktestConfig(
        ...     start_date="2020-01-01",
        ...     end_date="2023-12-31",
        ...     initial_capital=100000
        ... )
        >>> backtester = Backtester(config)
        >>> results = backtester.run_backtest(["AAPL", "MSFT", "GOOGL"])
    """
    
    RISK_FREE_RATE = 0.04  # 4% annual risk-free rate
    TRADING_DAYS_PER_YEAR = 252
    
    def __init__(self, config: BacktestConfig):
        """
        Initialize backtester with configuration.
        
        Args:
            config: BacktestConfig object with parameters
        """
        self.config = config
        self._price_cache: dict = {}
        self._data_cache: dict = {}
    
    def _fetch_historical_prices(self, ticker: str) -> Optional[pd.DataFrame]:
        """Fetch historical prices for a ticker"""
        if ticker in self._price_cache:
            return self._price_cache[ticker]
        
        try:
            stock = yf.Ticker(ticker)
            df = stock.history(
                start=self.config.start_date,
                end=self.config.end_date,
                interval="1d"
            )
            if len(df) > 0:
                self._price_cache[ticker] = df
                return df
        except Exception as e:
            print(f"Error fetching {ticker}: {e}")
        return None
    
    def _fetch_stock_data(self, ticker: str, date: datetime) -> Optional[dict]:
        """Fetch stock fundamental data for a specific date"""
        cache_key = f"{ticker}_{date.strftime('%Y-%m')}"
        if cache_key in self._data_cache:
            return self._data_cache[cache_key]
        
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            
            if not info or not info.get("symbol"):
                return None
            
            # Get historical price for context
            prices = self._fetch_historical_prices(ticker)
            
            data = {
                "symbol": ticker,
                "name": info.get("shortName") or info.get("longName") or "N/A",
                "sector": info.get("sector") or "N/A",
                "industry": info.get("industry") or "N/A",
                "price": info.get("currentPrice") or info.get("regularMarketPrice"),
                "market_cap": info.get("marketCap"),
                "pe_ratio": info.get("trailingPE"),
                "pb_ratio": info.get("priceToBook"),
                "dividend_yield": info.get("dividendYield"),
                "eps": info.get("trailingEps"),
                "beta": info.get("beta"),
                "52_week_high": info.get("fiftyTwoWeekHigh"),
                "52_week_low": info.get("fiftyTwoWeekLow"),
                "50_day_avg": info.get("fiftyDayAverage"),
                "200_day_avg": info.get("twoHundredDayAverage"),
                "revenue_growth": info.get("revenueGrowth"),
                "earnings_growth": info.get("earningsGrowth"),
                "roe": info.get("returnOnEquity"),
                "roa": info.get("returnOnAssets"),
                "debt_to_equity": info.get("debtToEquity"),
                "current_ratio": info.get("currentRatio"),
                "free_cash_flow": info.get("freeCashflow"),
                "operating_cash_flow": info.get("operatingCashflow"),
                "profit_margin": info.get("profitMargins"),
                "payout_ratio": info.get("payoutRatio"),
            }
            
            # Add historical context from prices
            if prices is not None and len(prices) > 0:
                close = prices["Close"]
                data["year_ago_price"] = float(close.iloc[0]) if len(close) > 0 else None
                if len(close) > 126:
                    data["6_month_ago_price"] = float(close.iloc[len(close)//2])
                if len(close) > 189:
                    data["3_month_ago_price"] = float(close.iloc[len(close)*3//4])
            
            self._data_cache[cache_key] = data
            return data
            
        except Exception as e:
            print(f"Error fetching data for {ticker}: {e}")
            return None
    
    def _get_rebalance_dates(self) -> list:
        """Generate list of rebalance dates based on frequency"""
        start = datetime.strptime(self.config.start_date, "%Y-%m-%d")
        end = datetime.strptime(self.config.end_date, "%Y-%m-%d")
        dates = [start]
        
        freq = self.config.rebalance_frequency
        
        if freq == RebalanceFrequency.DAILY:
            current = start
            while current < end:
                current += timedelta(days=1)
                if current.weekday() < 5:  # Weekdays only
                    dates.append(current)
        
        elif freq == RebalanceFrequency.WEEKLY:
            current = start
            while current < end:
                current += timedelta(weeks=1)
                dates.append(current)
        
        elif freq == RebalanceFrequency.MONTHLY:
            current = start
            while current < end:
                # Move to next month
                if current.month == 12:
                    current = current.replace(year=current.year + 1, month=1)
                else:
                    current = current.replace(month=current.month + 1)
                dates.append(current)
        
        elif freq == RebalanceFrequency.QUARTERLY:
            current = start
            while current < end:
                # Move to next quarter
                new_month = current.month + 3
                new_year = current.year
                if new_month > 12:
                    new_month -= 12
                    new_year += 1
                current = current.replace(year=new_year, month=new_month)
                dates.append(current)
        
        elif freq == RebalanceFrequency.YEARLY:
            current = start
            while current < end:
                current = current.replace(year=current.year + 1)
                dates.append(current)
        
        return [d for d in dates if d <= end]
    
    def _select_stocks(self, tickers: list, date: datetime) -> list:
        """Select stocks to buy based on strategy scores"""
        selected = []
        
        for ticker in tickers:
            data = self._fetch_stock_data(ticker, date)
            if data is None:
                continue
            
            analysis = InvestmentStrategies.analyze_stock(data)
            total_score = analysis.get("total_score", 0)
            
            if total_score >= self.config.min_score_threshold:
                selected.append({
                    "ticker": ticker,
                    "score": total_score,
                    "data": data,
                    "analysis": analysis
                })
        
        # Sort by score descending
        selected.sort(key=lambda x: x["score"], reverse=True)
        return selected
    
    def _calculate_metrics(self, portfolio_values: pd.DataFrame, 
                          benchmark_values: pd.DataFrame) -> dict:
        """Calculate performance metrics"""
        if len(portfolio_values) < 2:
            return {}
        
        # Daily returns
        portfolio_returns = portfolio_values.pct_change().dropna()
        benchmark_returns = benchmark_values.pct_change().dropna()
        
        # Align returns
        aligned = portfolio_returns.to_frame("portfolio").join(
            benchmark_returns.to_frame("benchmark"), how="inner"
        )
        
        if len(aligned) < 2:
            return {}
        
        port_ret = aligned["portfolio"]
        bench_ret = aligned["benchmark"]
        
        # Total return
        total_return = (portfolio_values.iloc[-1] / portfolio_values.iloc[0]) - 1
        
        # Annualized return
        days = (portfolio_values.index[-1] - portfolio_values.index[0]).days
        years = max(days / 365.25, 0.1)
        annualized_return = (1 + total_return) ** (1 / years) - 1
        
        # Benchmark return
        bench_total_return = (benchmark_values.iloc[-1] / benchmark_values.iloc[0]) - 1
        
        # Volatility (annualized)
        volatility = port_ret.std() * np.sqrt(self.TRADING_DAYS_PER_YEAR)
        
        # Sharpe Ratio
        excess_returns = port_ret - (self.RISK_FREE_RATE / self.TRADING_DAYS_PER_YEAR)
        sharpe_ratio = np.sqrt(self.TRADING_DAYS_PER_YEAR) * excess_returns.mean() / port_ret.std() if port_ret.std() > 0 else 0
        
        # Sortino Ratio (downside deviation)
        downside_returns = port_ret[port_ret < 0]
        downside_std = downside_returns.std() if len(downside_returns) > 0 else 0
        sortino_ratio = np.sqrt(self.TRADING_DAYS_PER_YEAR) * excess_returns.mean() / downside_std if downside_std > 0 else 0
        
        # Maximum Drawdown
        rolling_max = portfolio_values.cummax()
        drawdown = (portfolio_values - rolling_max) / rolling_max
        max_drawdown = abs(drawdown.min())
        
        # Max drawdown duration
        in_drawdown = drawdown < 0
        drawdown_periods = []
        start_idx = None
        for i, (idx, is_dd) in enumerate(in_drawdown.items()):
            if is_dd and start_idx is None:
                start_idx = i
            elif not is_dd and start_idx is not None:
                drawdown_periods.append(i - start_idx)
                start_idx = None
        if start_idx is not None:
            drawdown_periods.append(len(in_drawdown) - start_idx)
        max_drawdown_duration = max(drawdown_periods) if drawdown_periods else 0
        
        # Beta and Alpha
        if bench_ret.std() > 0:
            covariance = port_ret.cov(bench_ret)
            beta = covariance / (bench_ret.var())
            alpha = annualized_return - (self.RISK_FREE_RATE + beta * (bench_total_return / years - self.RISK_FREE_RATE))
        else:
            beta = 1.0
            alpha = 0.0
        
        return {
            "total_return": total_return,
            "annualized_return": annualized_return,
            "benchmark_return": bench_total_return,
            "volatility": volatility,
            "sharpe_ratio": sharpe_ratio,
            "sortino_ratio": sortino_ratio,
            "max_drawdown": max_drawdown,
            "max_drawdown_duration": max_drawdown_duration,
            "beta": beta,
            "alpha": alpha,
            "portfolio_returns": port_ret,
            "benchmark_returns": bench_ret,
        }
    
    def run_backtest(self, tickers: list) -> BacktestResults:
        """
        Run backtest on given tickers.
        
        Args:
            tickers: List of ticker symbols to backtest
            
        Returns:
            BacktestResults object with all metrics and data
        """
        print(f"\n{'='*60}")
        print(f"Starting Backtest")
        print(f"{'='*60}")
        print(f"Period: {self.config.start_date} to {self.config.end_date}")
        print(f"Initial Capital: ${self.config.initial_capital:,.2f}")
        print(f"Rebalance: {self.config.rebalance_frequency.value}")
        print(f"Tickers: {len(tickers)}")
        print(f"{'='*60}\n")
        
        # Initialize
        capital = self.config.initial_capital
        positions: dict[str, Position] = {}
        trades: list[Trade] = []
        rebalance_dates = self._get_rebalance_dates()
        
        # Track portfolio values daily
        all_dates = pd.date_range(
            start=self.config.start_date,
            end=self.config.end_date,
            freq="B"  # Business days
        )
        portfolio_values = pd.Series(index=all_dates, dtype=float)
        
        # Fetch benchmark data
        benchmark_df = self._fetch_historical_prices(self.config.benchmark)
        if benchmark_df is None:
            print(f"Warning: Could not fetch benchmark ({self.config.benchmark})")
            benchmark_df = pd.DataFrame()
        
        print(f"Running backtest with {len(rebalance_dates)} rebalance dates...")
        
        # Main backtest loop
        for rebal_idx, rebal_date in enumerate(rebalance_dates):
            # Update position prices for this date
            for ticker, pos in positions.items():
                prices = self._fetch_historical_prices(ticker)
                if prices is not None:
                    # Find closest price date
                    mask = prices.index <= rebal_date
                    if mask.any():
                        pos.current_price = float(prices.loc[mask, "Close"].iloc[-1])
            
            # Calculate current portfolio value
            position_values = {t: p.market_value for t, p in positions.items()}
            total_position_value = sum(position_values.values())
            portfolio_value = capital + total_position_value
            
            # Store portfolio value
            try:
                portfolio_values[rebal_date] = portfolio_value
            except KeyError:
                pass
            
            # Sell logic: stocks that no longer meet threshold
            stocks_to_sell = []
            for ticker, pos in positions.items():
                data = self._fetch_stock_data(ticker, rebal_date)
                if data:
                    analysis = InvestmentStrategies.analyze_stock(data)
                    if analysis.get("total_score", 0) < self.config.min_score_threshold:
                        stocks_to_sell.append(ticker)
            
            # Execute sells
            for ticker in stocks_to_sell:
                pos = positions.pop(ticker)
                sell_value = pos.market_value
                transaction_cost = sell_value * self.config.transaction_cost_pct
                
                trades.append(Trade(
                    date=rebal_date,
                    ticker=ticker,
                    action="SELL",
                    shares=pos.shares,
                    price=pos.current_price,
                    value=sell_value,
                    transaction_cost=transaction_cost,
                    reason="Score below threshold"
                ))
                
                capital += sell_value - transaction_cost
                total_position_value -= sell_value
            
            # Rebalance: allocate capital to top-scoring stocks
            if rebal_idx < len(rebalance_dates) - 1:  # Not the last date
                selected_stocks = self._select_stocks(tickers, rebal_date)
                
                # Calculate target allocation
                n_positions = min(len(selected_stocks), 10)  # Max 10 positions
                if n_positions > 0:
                    available_capital = capital + total_position_value
                    target_per_position = available_capital * self.config.position_size_pct
                    
                    # Buy new positions
                    for stock in selected_stocks[:n_positions]:
                        ticker = stock["ticker"]
                        if ticker in positions:
                            continue  # Already held
                            
                        prices = self._fetch_historical_prices(ticker)
                        if prices is None:
                            continue
                            
                        mask = prices.index <= rebal_date
                        if not mask.any():
                            continue
                            
                        price = float(prices.loc[mask, "Close"].iloc[-1])
                        
                        # Calculate shares to buy
                        buy_value = min(target_per_position, available_capital * 0.95)
                        shares = int(buy_value / price)
                        
                        if shares > 0:
                            buy_cost = shares * price
                            transaction_cost = buy_cost * self.config.transaction_cost_pct
                            
                            trades.append(Trade(
                                date=rebal_date,
                                ticker=ticker,
                                action="BUY",
                                shares=shares,
                                price=price,
                                value=buy_cost,
                                transaction_cost=transaction_cost,
                                reason=f"Score: {stock['score']}/100"
                            ))
                            
                            positions[ticker] = Position(
                                ticker=ticker,
                                shares=shares,
                                avg_cost=price,
                                current_price=price
                            )
                            
                            capital -= (buy_cost + transaction_cost)
            
            if (rebal_idx + 1) % 5 == 0 or rebal_idx == len(rebalance_dates) - 1:
                print(f"  Rebalance {rebal_idx + 1}/{len(rebalance_dates)}: "
                      f"{len(positions)} positions, "
                      f"Capital: ${capital:,.2f}")
        
        # Final valuation
        final_date = datetime.strptime(self.config.end_date, "%Y-%m-%d")
        for ticker, pos in positions.items():
            prices = self._fetch_historical_prices(ticker)
            if prices is not None:
                mask = prices.index <= final_date
                if mask.any():
                    pos.current_price = float(prices.loc[mask, "Close"].iloc[-1])
        
        final_position_value = sum(p.market_value for p in positions.values())
        final_capital = capital + final_position_value
        
        # Fill remaining portfolio values
        for date in all_dates:
            if pd.isna(portfolio_values.get(date)):
                # Forward fill from last known value
                prev_dates = [d for d in portfolio_values.index if d < date]
                if prev_dates:
                    portfolio_values[date] = portfolio_values[prev_dates[-1]]
        
        # Calculate metrics
        if len(benchmark_df) > 0:
            benchmark_values = benchmark_df["Close"]
            benchmark_values = benchmark_values.reindex(all_dates, method="ffill")
        else:
            benchmark_values = pd.Series(100, index=all_dates)
        
        metrics = self._calculate_metrics(portfolio_values, benchmark_values)
        
        # Trade statistics
        buy_trades = [t for t in trades if t.action == "BUY"]
        sell_trades = [t for t in trades if t.action == "SELL"]
        winning_trades = sum(1 for t in sell_trades if t.price > 0)  # Simplified
        total_transaction_costs = sum(t.transaction_cost for t in trades)
        
        print(f"\n{'='*60}")
        print(f"Backtest Complete")
        print(f"{'='*60}")
        print(f"Final Capital: ${final_capital:,.2f}")
        print(f"Total Return: {(final_capital/self.config.initial_capital - 1):.2%}")
        print(f"Total Trades: {len(trades)}")
        print(f"{'='*60}\n")
        
        return BacktestResults(
            config=self.config,
            tickers=tickers,
            total_return=metrics.get("total_return", 0),
            annualized_return=metrics.get("annualized_return", 0),
            benchmark_return=metrics.get("benchmark_return", 0),
            excess_return=metrics.get("total_return", 0) - metrics.get("benchmark_return", 0),
            volatility=metrics.get("volatility", 0),
            sharpe_ratio=metrics.get("sharpe_ratio", 0),
            sortino_ratio=metrics.get("sortino_ratio", 0),
            max_drawdown=metrics.get("max_drawdown", 0),
            max_drawdown_duration=metrics.get("max_drawdown_duration", 0),
            beta=metrics.get("beta", 1.0),
            alpha=metrics.get("alpha", 0),
            total_trades=len(trades),
            winning_trades=winning_trades,
            losing_trades=len(sell_trades) - winning_trades,
            win_rate=winning_trades / len(sell_trades) if sell_trades else 0,
            total_transaction_costs=total_transaction_costs,
            portfolio_values=portfolio_values.to_frame("portfolio_value"),
            daily_returns=metrics.get("portfolio_returns", pd.Series()),
            benchmark_returns=metrics.get("benchmark_returns", pd.Series()),
            trades=trades,
            final_capital=final_capital,
            start_date=self.config.start_date,
            end_date=self.config.end_date,
        )
    
    def export_results(self, results: BacktestResults, output_dir: str = "output") -> str:
        """Export backtest results to Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = output_path / f"backtest_results_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = results.to_dict()
        row = 1
        for key, value in summary_data.items():
            ws_summary[f"A{row}"] = key.replace("_", " ").title()
            ws_summary[f"B{row}"] = f"{value:,.4f}" if isinstance(value, float) else value
            ws_summary[f"A{row}"].font = Font(bold=True)
            row += 1
        
        # Trades sheet
        ws_trades = wb.create_sheet("Trades")
        if results.trades:
            headers = ["Date", "Ticker", "Action", "Shares", "Price", "Value", "Cost", "Reason"]
            ws_trades.append(headers)
            for trade in results.trades:
                ws_trades.append([
                    trade.date.strftime("%Y-%m-%d"),
                    trade.ticker,
                    trade.action,
                    trade.shares,
                    f"${trade.price:.2f}",
                    f"${trade.value:,.2f}",
                    f"${trade.transaction_cost:.2f}",
                    trade.reason
                ])
        
        # Portfolio values sheet
        ws_values = wb.create_sheet("Portfolio Values")
        ws_values.append(["Date", "Portfolio Value", "Daily Return"])
        for date, value in results.portfolio_values.iterrows():
            daily_ret = results.portfolio_values.pct_change().loc[date] if len(results.portfolio_values) > 1 else 0
            ws_values.append([
                date.strftime("%Y-%m-%d"),
                f"${value['portfolio_value']:,.2f}",
                f"{daily_ret:.4%}" if not pd.isna(daily_ret) else "N/A"
            ])
        
        # Save
        wb.save(filepath)
        print(f"Results exported to: {filepath}")
        
        return str(filepath)


if __name__ == "__main__":
    # Example backtest
    config = BacktestConfig(
        start_date="2022-01-01",
        end_date="2023-12-31",
        initial_capital=100000,
        rebalance_frequency="monthly"
    )
    
    backtester = Backtester(config)
    tickers = ["AAPL", "MSFT", "GOOGL", "AMZN", "META", "NVDA", "TSLA", "JPM", "JNJ", "WMT"]
    
    results = backtester.run_backtest(tickers)
    print(f"\nResults Summary:")
    print(f"  Total Return: {results.total_return:.2%}")
    print(f"  Sharpe Ratio: {results.sharpe_ratio:.2f}")
    print(f"  Max Drawdown: {results.max_drawdown:.2%}")
