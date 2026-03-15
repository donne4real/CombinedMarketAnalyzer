"""
Spreadsheet Export Module
Exports stock analysis results to Excel format.
"""

import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Color
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

from .strategies import STRATEGY_NAMES


class SpreadsheetExporter:
    """Exports stock analysis results to formatted Excel spreadsheets."""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALTERNATE_ROW = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def create_dataframe(self, stocks_data: list, strategies_data: list) -> pd.DataFrame:
        """
        Create a pandas DataFrame combining stock data and strategy scores.
        """
        rows = []

        for stock, strategies in zip(stocks_data, strategies_data):
            row = {
                "Ticker": stock.get("symbol", ""),
                "Company Name": stock.get("name", ""),
                "Sector": stock.get("sector", ""),
                "Industry": stock.get("industry", ""),
                "Price ($)": stock.get("price"),
                "Market Cap": stock.get("market_cap"),
                "P/E Ratio": stock.get("pe_ratio"),
                "P/B Ratio": stock.get("pb_ratio"),
                "Dividend Yield": stock.get("dividend_yield"),
                "Beta": stock.get("beta"),
                "ROE": stock.get("roe"),
                "Revenue Growth": stock.get("revenue_growth"),
            }

            # Add strategy scores
            for strategy_key, strategy_name in STRATEGY_NAMES.items():
                strategy_result = strategies.get(strategy_key, {})
                row[strategy_name] = strategy_result.get("score", 0)

            # Add totals
            row["Total Score"] = strategies.get("total_score", 0)
            row["Average Score"] = strategies.get("average_score", 0)
            row["Analysis Notes"] = "; ".join([
                strategies.get(k, {}).get("reason", "")
                for k in STRATEGY_NAMES.keys()
                if strategies.get(k, {}).get("reason")
            ][:500])  # Limit length

            rows.append(row)

        return pd.DataFrame(rows)

    def export_to_excel(
        self,
        stocks_data: list,
        strategies_data: list,
        filename: str = None
    ) -> str:
        """
        Export analysis results to a formatted Excel file.
        Returns the path to the created file.
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"stock_analysis_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Create DataFrame
        df = self.create_dataframe(stocks_data, strategies_data)

        # Sort by total score (descending)
        df = df.sort_values("Total Score", ascending=False)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Analysis"

        # Write headers
        headers = df.columns.tolist()
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER

        # Write data rows
        for idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=idx, column=col_idx, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Alternate row colors
                if idx % 2 == 0:
                    cell.fill = self.ALTERNATE_ROW

                # Center-align numeric columns
                if col_idx in [5, 6, 7, 8, 9, 10, 11, 12, 13]:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                # Format strategy score columns (green scale)
                if col_idx > 13 and col_idx <= 23:  # Strategy scores
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if value >= 8:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif value >= 6:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            cell.font = Font(color="9C5700")
                        elif value <= 3:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")

        # Auto-adjust column widths
        column_widths = {
            "Ticker": 10,
            "Company Name": 35,
            "Sector": 18,
            "Industry": 25,
            "Price ($)": 12,
            "Market Cap": 15,
            "P/E Ratio": 12,
            "P/B Ratio": 12,
            "Dividend Yield": 15,
            "Beta": 8,
            "ROE": 10,
            "Revenue Growth": 15,
        }

        # Strategy columns width
        for strategy_name in STRATEGY_NAMES.values():
            column_widths[strategy_name] = 12

        column_widths["Total Score"] = 12
        column_widths["Average Score"] = 15
        column_widths["Analysis Notes"] = 50

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            width = column_widths.get(header, 15)
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add summary sheet
        self._add_summary_sheet(wb, stocks_data, strategies_data)

        # Save workbook
        wb.save(filepath)
        print(f"\n✓ Exported to: {filepath}")

        return str(filepath)

    def _add_summary_sheet(self, wb: Workbook, stocks_data: list, strategies_data: list):
        """Add a summary sheet with aggregate statistics."""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Stock Analysis Summary"
        ws["A1"].font = Font(bold=True, size=16)

        # Basic stats
        stats = [
            ("Total Stocks Analyzed", len(stocks_data)),
            ("Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Top Strategies (by avg score)", ""),
        ]

        # Calculate average score per strategy
        strategy_avgs = {}
        for strategy_key in STRATEGY_NAMES.keys():
            scores = [s.get(strategy_key, {}).get("score", 0) for s in strategies_data]
            if scores:
                strategy_avgs[strategy_key] = sum(scores) / len(scores)

        # Sort strategies by average score
        sorted_strategies = sorted(strategy_avgs.items(), key=lambda x: x[1], reverse=True)

        for strategy_key, avg_score in sorted_strategies:
            stats.append((STRATEGY_NAMES[strategy_key], f"{avg_score:.2f}"))

        # Top 10 stocks
        stats.append(("", ""))
        stats.append(("Top 10 Stocks (by total score)", ""))

        # Get top 10 stocks
        stock_scores = []
        for stock, strategies in zip(stocks_data, strategies_data):
            stock_scores.append((
                stock.get("symbol", ""),
                stock.get("name", ""),
                strategies.get("total_score", 0),
                strategies.get("average_score", 0)
            ))

        stock_scores.sort(key=lambda x: x[2], reverse=True)

        for symbol, name, total, avg in stock_scores[:10]:
            stats.append((symbol, f"{name[:30]}... | Total: {total}, Avg: {avg}"))

        # Write stats
        for row_idx, (label, value) in enumerate(stats, start=3):
            ws[f"A{row_idx}"] = label
            ws[f"B{row_idx}"] = value
            ws[f"A{row_idx}"].font = Font(bold=True) if label else Font()

        # Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 55

    def export_to_csv(
        self,
        stocks_data: list,
        strategies_data: list,
        filename: str = None
    ) -> str:
        """Export analysis results to CSV format."""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"stock_analysis_{timestamp}.csv"

        filepath = self.output_dir / filename

        df = self.create_dataframe(stocks_data, strategies_data)
        df = df.sort_values("Total Score", ascending=False)
        df.to_csv(filepath, index=False)

        print(f"\n✓ Exported to: {filepath}")
        return str(filepath)


if __name__ == "__main__":
    # Test export
    test_stocks = [
        {
            "symbol": "AAPL",
            "name": "Apple Inc.",
            "sector": "Technology",
            "industry": "Consumer Electronics",
            "price": 175.50,
            "market_cap": 2800000000000,
            "pe_ratio": 28.5,
            "pb_ratio": 45.2,
            "dividend_yield": 0.005,
            "beta": 1.2,
            "roe": 0.15,
            "revenue_growth": 0.08,
        }
    ]

    test_strategies = [
        {
            "value": {"score": 6, "reason": "Moderate P/E"},
            "growth": {"score": 7, "reason": "Good growth"},
            "dividend": {"score": 2, "reason": "Low yield"},
            "momentum": {"score": 8, "reason": "Strong momentum"},
            "quality": {"score": 9, "reason": "High ROE"},
            "small_cap": {"score": 0, "reason": "Large cap"},
            "large_cap": {"score": 10, "reason": "Mega cap"},
            "low_volatility": {"score": 5, "reason": "Market beta"},
            "peg_ratio": {"score": 6, "reason": "Fair PEG"},
            "free_cash_flow": {"score": 8, "reason": "Strong FCF"},
            "total_score": 61,
            "average_score": 6.1,
        }
    ]

    exporter = SpreadsheetExporter()
    exporter.export_to_excel(test_stocks, test_strategies)
